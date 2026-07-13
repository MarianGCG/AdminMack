from django.shortcuts import render

from .models import Movimiento, Categoria, Finalidad, Persona,  Regla
from comisiones.models import ParametroSistema
from django.http import JsonResponse
import json
from django.views.decorators.http import require_GET
from .services.movimientos_service import importar_movimientos,  buscar_regla, actualizar_movimientos
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Sum, Avg
from django.db.models.functions import Coalesce
from django.db.models import DecimalField
from django.contrib import messages

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import Movimiento

from django.core.serializers.json import DjangoJSONEncoder
from django.contrib.auth.decorators import login_required
from django.db import DataError, IntegrityError

@require_GET
def categoria_eliminar(request, id):

    Categoria.objects.filter(id=id).delete()

    return JsonResponse({"ok": True})




@login_required
def categorias(request):
    categorias = Categoria.objects.order_by("nombre")

    return render(
        request,
        "finanzas/categorias.html",
        {
            "categorias": categorias
        }
    )

def categoria_guardar(request):

    if request.method != "POST":
        return JsonResponse({"ok": False})

    datos = json.loads(request.body)

    id = datos.get("id")

    if id:
        categoria = Categoria.objects.get(id=id)
    else:
        categoria = Categoria()

    categoria.codigo = datos.get("codigo")
    categoria.nombre = datos.get("nombre")
    categoria.color = datos.get("color")
    categoria.activo = datos.get("activo")

    try:

        categoria.save()

        return JsonResponse({
            "ok": True
        })

    except DataError:

        return JsonResponse({
            "ok": False,
            "error": "El código es demasiado largo."
        })

    except IntegrityError:

        return JsonResponse({
            "ok": False,
            "error": "Ya existe una categoría con ese código."
        })

    except Exception as e:

        return JsonResponse({
            "ok": False,
            "error": str(e)
        })

@login_required
def finalidades(request):

    finalidades = Finalidad.objects.order_by("nombre")

    return render(
        request,
        "finanzas/finalidades.html",
        {
            "finalidades": finalidades
        }
    )

def finalidad_guardar(request):

    if request.method != "POST":
        return JsonResponse({"ok": False})

    datos = json.loads(request.body)

    id = datos.get("id")

    if id:
        finalidad = Finalidad.objects.get(id=id)
    else:
        finalidad = Finalidad()

    finalidad.codigo = datos.get("codigo")
    finalidad.nombre = datos.get("nombre")
    finalidad.activo = datos.get("activo")

    finalidad.save()

    return JsonResponse({"ok": True})

def finalidad_eliminar(request, id):

    Finalidad.objects.filter(id=id).delete()

    return JsonResponse({"ok": True})




@login_required
def personas(request):

    personas = Persona.objects.order_by("nombre")

    return render(
        request,
        "finanzas/personas.html",
        {
            "personas": personas
        }
    )

def persona_guardar(request):

    if request.method != "POST":
        return JsonResponse({"ok": False})

    datos = json.loads(request.body)

    id = datos.get("id")

    if id:
        persona = Persona.objects.get(id=id)
    else:
        persona = Persona()

    persona.codigo = datos.get("codigo")
    persona.nombre = datos.get("nombre")
    persona.activo = datos.get("activo")
    persona.save()

    return JsonResponse({"ok": True})



def persona_eliminar(request, id):

    Persona.objects.filter(id=id).delete()

    return JsonResponse({"ok": True})





# ==========================================
# REGLAS
# ==========================================



@login_required
def reglas(request):

    reglas = Regla.objects.order_by("texto")

    popup = request.GET.get("popup") == "1"
    crear = request.GET.get("crear") == "1"
    texto = request.GET.get("texto", "")

    return render(
        request,
        "finanzas/reglas.html",
        {
            "reglas": reglas,
            "categorias": Categoria.objects.order_by("nombre"),
            "finalidades": Finalidad.objects.order_by("nombre"),
            "personas": Persona.objects.order_by("nombre"),
            "parametros": ParametroSistema.objects.filter(valor="REGLA").order_by("codigo"),
            "popup": popup,
            "crear": crear,
            "texto_inicial": texto,
        }
    )


def regla_guardar(request):

    print("******** ENTRO A REGLA_GUARDAR ********")

    if request.method != "POST":
        return JsonResponse({"ok": False})

    datos = json.loads(request.body)
    print(datos)

    id = datos.get("id")
    es_nueva = False

    if id:

        regla = Regla.objects.get(id=id)

    else:

        regla = Regla.objects.filter(
            texto=datos.get("texto")
        ).first()

        if regla is None:
            regla = Regla()
            es_nueva = True



    regla.texto = datos.get("texto")
    regla.accion = datos.get("accion")

    regla.categoria_id = datos.get("categoria") or None
    regla.finalidad_id = datos.get("finalidad") or None
    regla.persona_id = datos.get("persona") or None

    regla.grupo = datos.get("grupo", "")

    regla.activa = datos.get("activo")
    regla.observacion = datos.get("observacion", "")
    print(datos)

    print("Categoria:", datos.get("categoria"))
    print("Finalidad:", datos.get("finalidad"))
    print("Persona:", datos.get("persona"))
    regla.save()

    # ---------------------------------
    # Reaplicar reglas automáticamente
    # ---------------------------------
    actualizar_movimientos()

    return JsonResponse({
        "ok": True,
        "nueva": es_nueva
    })
    
def regla_eliminar(request, id):

    Regla.objects.filter(id=id).delete()

    return JsonResponse({"ok": True})


# ==========================================
# MOVIMIENTOS
# ==========================================
@login_required
def movimientos(request):

    movimientos = Movimiento.objects.select_related(
        "categoria",
        "finalidad",
        "persona",
        "regla_aplicada",
    )

    # ------------------------
    # FILTROS
    # ------------------------

    periodo = request.GET.get("periodo", "")
    if periodo:
        movimientos = movimientos.filter(periodo=periodo)
        
    categoria = request.GET.get("categoria")

    if categoria:
        movimientos = movimientos.filter(
            categoria_id=categoria
        )

    finalidad = request.GET.get("finalidad")

    if finalidad:
        movimientos = movimientos.filter(
            finalidad_id=finalidad
        )

    persona = request.GET.get("persona")

    if persona:
        movimientos = movimientos.filter(
            persona_id=persona
        )

  

    descripcion = request.GET.get("descripcion", "").strip()

    if descripcion:
        movimientos = movimientos.filter(
            descripcion__icontains=descripcion
        )






  
    # ---------------------------------------
    # PANEL ELIMINAR IMPORTACIÓN
    # ---------------------------------------

    mostrar_eliminar = request.GET.get("eliminar") == "1"

    archivos_importados = []

    if mostrar_eliminar:

        archivos_importados = (
            Movimiento.objects
            .exclude(nombre_archivo__isnull=True)
            .exclude(nombre_archivo="")
            .values("nombre_archivo")
            .annotate(
                cantidad=Count("id")
            )
            .order_by("-nombre_archivo")
        )
      

    # ------------------------
    # ORDEN
    # ------------------------

    orden = request.GET.get("orden")

    if orden:
        movimientos = movimientos.order_by(orden)
    else:
        orden = "-periodo"
        movimientos = movimientos.order_by(
            "-periodo",
            "-regla_aplicada",
            "-fecha"
        )

    periodos = (
        Movimiento.objects
        .exclude(periodo="")
        .values("periodo")
        .distinct()
        .order_by("-periodo")
    )



    # =====================================
    # RESUMEN POR PERIODO / ORIGEN
    # =====================================

    resumen_origen = (
        movimientos
        .values("periodo", "origen")
        .annotate(
            cantidad=Count("id"),
            total=Sum("importe")
        )
        .order_by("periodo", "origen")
    )



    totales = movimientos.aggregate(
        cantidad=Count("id"),
        total=Sum("importe")
    )

    return render(
        request,
        "finanzas/movimientos.html",
        {

            "movimientos": movimientos,


            "categorias": Categoria.objects.order_by("nombre"),
            "finalidades": Finalidad.objects.order_by("nombre"),
            "personas": Persona.objects.order_by("nombre"),

            "periodos": periodos,
            "orden": orden,

            # NUEVO
            "resumen_origen": resumen_origen,
            "totales": totales,
            "mostrar_eliminar": mostrar_eliminar,
            "archivos_importados": archivos_importados,

        }
    )

def movimiento_guardar(request):

    return JsonResponse({"ok": True})


def movimiento_eliminar(request, id):

    return JsonResponse({"ok": True})


def movimiento_importar(request):

    if request.method != "POST":
        return JsonResponse({
            "ok": False,
            "error": "Método no permitido"
        })

    archivo = request.FILES.get("archivo")

    if not archivo:
        return JsonResponse({
            "ok": False,
            "error": "No se recibió ningún archivo"
        })

    try:

        mensaje = importar_movimientos(archivo)

        return JsonResponse({
            "ok": True,
            "mensaje": mensaje
        })

    except Exception as e:

        return JsonResponse({
            "ok": False,
            "error": str(e)
        })




def movimiento_actualizar(request):

    actualizar_movimientos()

    return redirect("movimientos")

#############################################################################################################
# DASHBOARD
#############################################################################################################
@login_required
def dashboard(request):

    periodo_actual = request.GET.get("periodo")

    if not periodo_actual:

        periodo_actual = (
            Movimiento.objects
            .order_by("-periodo")
            .values_list("periodo", flat=True)
            .first()
        )

    movimientos = Movimiento.objects.filter(
        periodo=periodo_actual
    )

    #==========================
    # KPIs
    #==========================

    total_gastos = movimientos.aggregate(

        total=Coalesce(

            Sum("importe"),

            0,

            output_field=DecimalField()

        )

    )["total"]


    cantidad_movimientos = movimientos.count()


    ticket_promedio = movimientos.aggregate(

        promedio=Coalesce(

            Avg("importe"),

            0,

            output_field=DecimalField()

        )

    )["promedio"]


    cantidad_categorias = (
        movimientos
        .values("categoria")
        .distinct()
        .count()
    )



    #==========================
    # GASTOS POR CATEGORIA
    #==========================

    gastos_categoria = (
        movimientos
        .values("categoria__nombre")
        .annotate(
            total=Sum("importe")
        )
        .order_by("-total")
    )

    categoria_principal = gastos_categoria.first()

    #==========================
    # PERSONAS
    #==========================

    gastos_persona = (

        movimientos

        .values("persona__nombre")

        .annotate(

            total=Sum("importe")

        )

        .order_by("-total")

    )



    #==========================
    # FINALIDAD
    #==========================

    gastos_finalidad = (

        movimientos

        .values("finalidad__nombre")

        .annotate(

            total=Sum("importe")

        )

        .order_by("-total")

    )



    #==========================
    # ORIGEN
    #==========================

    gastos_origen = (

        movimientos

        .values("origen")

        .annotate(

            total=Sum("importe")

        )

        .order_by("-total")

        )

    ultimos_movimientos = movimientos[:20]

    datos_dashboard = []

    for m in movimientos.select_related(
        "categoria",
        "finalidad",
        "persona",
    ):

        datos_dashboard.append({

            "fecha": m.fecha,
            "periodo": m.periodo,
            "descripcion": m.descripcion,
            "importe": float(m.importe),
            "categoria": m.categoria.nombre if m.categoria else "Sin clasificar",
            "finalidad": m.finalidad.nombre if m.finalidad else "Sin clasificar",
            "persona": m.persona.nombre if m.persona else "Sin asignar",
            "origen": m.origen,
            "archivo": m.nombre_archivo,
            "tipo": m.get_tipo_display(),

        })
        



    periodos = (

        Movimiento.objects

        .values("periodo")

        .distinct()

        .order_by("-periodo")

    )



    contexto = {

        "periodo_actual": periodo_actual,
        "periodos": periodos,
        "total_gastos": total_gastos,
        "cantidad_movimientos": cantidad_movimientos,
        "ticket_promedio": ticket_promedio,
        "cantidad_categorias": cantidad_categorias,
        "ultimos_movimientos": ultimos_movimientos,
        "gastos_categoria": list(gastos_categoria),
        "categoria_principal": categoria_principal,
        "gastos_persona": list(gastos_persona),
        "gastos_finalidad": list(gastos_finalidad),
        "gastos_origen": list(gastos_origen),
        "datos_json": json.dumps(
            datos_dashboard,
            cls=DjangoJSONEncoder
        ),


    }

    return render(

        request,

        "finanzas/dashboard.html",

        contexto

    )

from django.db.models import Count

def eliminar_resumen(request):
    print("******** ENTRO A eliminar_resumen ********")

    if request.method == "POST":

        nombre_archivo = request.POST.get("nombre_archivo")

        cantidad = Movimiento.objects.filter(
            nombre_archivo=nombre_archivo
        ).count()

        Movimiento.objects.filter(
            nombre_archivo=nombre_archivo
        ).delete()

        messages.success(
            request,
            f"Se eliminaron {cantidad} movimientos."
        )

        return redirect("movimientos")

    # ==========================
    # Mostrar archivos
    # ==========================

    archivos = (
        Movimiento.objects
        .values("nombre_archivo")
        .annotate(cantidad=Count("id"))
        .order_by("-nombre_archivo")
    )

    return render(
        request,
        "finanzas/eliminar_resumen.html",
        {
            "archivos": archivos
        }
    )



@login_required
def movimientos_exportar_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    encabezados = [
        "Fecha",
        "Período",
        "Tipo",
        "Descripción",
        "Importe",
        "Categoría",
        "Finalidad",
        "Persona",
        "Origen",
        "Grupo",
        "Observación",
        "Archivo"
    ]

    ws.append(encabezados)

    # Encabezados en negrita
    for celda in ws[1]:
        celda.font = Font(bold=True)

    movimientos = (
        Movimiento.objects
        .select_related(
            "categoria",
            "finalidad",
            "persona"
        )
        .order_by("-fecha", "-id")
    )

    for m in movimientos:

        ws.append([
            m.fecha,
            m.periodo,
            m.get_tipo_display(),
            m.descripcion,
            float(m.importe),
            m.categoria.nombre if m.categoria else "",
            m.finalidad.nombre if m.finalidad else "",
            m.persona.nombre if m.persona else "",
            m.origen,
            m.grupo,
            m.observacion,
            m.nombre_archivo,
        ])

    # Formato de columnas
    for fila in ws.iter_rows(min_row=2):
        fila[0].number_format = "dd/mm/yyyy"   # Fecha
        fila[4].number_format = '#,##0.00'     # Importe

    # Ajustar ancho automáticamente
    for columna in ws.columns:
        largo = max(len(str(celda.value or "")) for celda in columna)
        ws.column_dimensions[get_column_letter(columna[0].column)].width = min(largo + 3, 50)

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Movimientos.xlsx"'

    wb.save(response)

    return response